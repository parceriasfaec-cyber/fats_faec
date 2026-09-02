import io
import os
import re
import sys
import uuid
from pathlib import Path

import psycopg2
import requests
from flask import (
    Flask, render_template, request, redirect, url_for, flash, send_file,
    send_from_directory, abort
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from database import get_connection, init_db, _base_dir
from pdf_generator import generate_pdf
from supabase_storage import enviar_bytes
from campos import FIELDS, FIELD_LABELS, MUNICIPIOS_CEARA
from fila_offline import (
    adicionar_na_fila, listar_pendentes, contar_pendentes, remover_da_fila,
    FILA_FOTOS_DIR,
)


# Pasta onde as fotos dos produtores ficam salvas (ao lado do banco de
# dados, para nao se perder quando o programa roda como .exe)
# Pasta onde as fotos dos produtores ficam salvas. Em producao (Render),
# aponte a variavel de ambiente FOTOS_DIR para o disco persistente
# (ex: /var/data/fotos). Localmente, sem essa variavel, continua usando a
# pasta "fotos" do lado do proprio app.py (como sempre foi).
UPLOAD_DIR = Path(os.environ["FOTOS_DIR"]) if os.environ.get("FOTOS_DIR") else (_base_dir() / "fotos")
try:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    # Em hospedagens com disco somente-leitura (Vercel, por exemplo), essa
    # pasta local nao pode ser criada - sem problema, pois as fotos novas
    # vao direto para o Supabase Storage. Essa pasta so serve para exibir
    # fotos antigas, salvas localmente antes da migracao.
    pass
EXTENSOES_PERMITIDAS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}


# Campos que devem ser sempre salvos em CAIXA ALTA
CAMPOS_MAIUSCULOS = {
    "nome_produtor", "nome_propriedade", "municipio",
    "tecnico_responsavel", "nome_tecnico",
}

# Campos que aceitam apenas numeros inteiros (sem letras nem simbolos)
CAMPOS_SOMENTE_NUMERO = {
    "cpf_tecnico",
    "membros_residentes", "vacas_lactacao", "vacas_secas",
    "novilhas", "touros", "qtd_receptoras",
}

# Campos numericos que podem ter casas decimais (usando virgula, padrao BR)
CAMPOS_DECIMAL = {
    "area_leite", "volume_diario", "produtividade_media",
    "capacidade_tanque", "estoque_seca", "area_palma", "area_capineira",
}


def _formatar_cpf(valor: str) -> str:
    """Aplica a mascara padrao de CPF (000.000.000-00), a partir de
    qualquer coisa que o usuario tenha digitado (com ou sem pontuacao)."""
    digitos = re.sub(r"\D", "", valor or "")[:11]
    if len(digitos) <= 3:
        return digitos
    if len(digitos) <= 6:
        return f"{digitos[0:3]}.{digitos[3:]}"
    if len(digitos) <= 9:
        return f"{digitos[0:3]}.{digitos[3:6]}.{digitos[6:]}"
    return f"{digitos[0:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"


def _formatar_telefone(valor: str) -> str:
    """Aplica mascara de telefone brasileiro: (00) 0000-0000 para fixo
    ou (00) 00000-0000 para celular, conforme a quantidade de digitos."""
    digitos = re.sub(r"\D", "", valor or "")[:11]
    if len(digitos) <= 2:
        return digitos
    if len(digitos) <= 6:
        return f"({digitos[0:2]}) {digitos[2:]}"
    if len(digitos) <= 10:
        return f"({digitos[0:2]}) {digitos[2:6]}-{digitos[6:]}"
    return f"({digitos[0:2]}) {digitos[2:7]}-{digitos[7:]}"


def _normalizar_data(valor: str) -> str:
    """Garante que a data de nascimento fique salva no formato dd/mm/aaaa."""
    valor = (valor or "").strip()
    if not valor:
        return ""
    # Ja esta no formato dd/mm/aaaa
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", valor):
        return valor
    # Formato vindo de <input type="date"> (aaaa-mm-dd)
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", valor)
    if m:
        aaaa, mm, dd = m.groups()
        return f"{dd}/{mm}/{aaaa}"
    # Usuario digitou so os numeros (ddmmaaaa)
    digitos = re.sub(r"\D", "", valor)
    if len(digitos) == 8:
        return f"{digitos[0:2]}/{digitos[2:4]}/{digitos[4:8]}"
    return valor


def _exibir_data(valor: str) -> str:
    """Converte datas antigas (aaaa-mm-dd) para dd/mm/aaaa na hora de exibir no formulario."""
    valor = (valor or "").strip()
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", valor)
    if m:
        aaaa, mm, dd = m.groups()
        return f"{dd}/{mm}/{aaaa}"
    return valor


def _foto_atual(pid) -> str:
    """Retorna o nome da foto ja salva para esse produtor (para nao perder
    a foto quando o formulario e reenviado sem escolher um novo arquivo)."""
    if pid is None:
        return ""
    conn = get_connection()
    row = conn.execute(
        "SELECT foto_produtor FROM produtores WHERE id = ?", (pid,)
    ).fetchone()
    conn.close()
    return (row["foto_produtor"] if row else "") or ""


def _resource_dir() -> Path:
    """Pasta onde estao templates/static, tanto rodando com 'python app.py'
    quanto rodando como .exe gerado pelo PyInstaller (modo --onefile)."""
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    return Path(__file__).parent


_BASE = _resource_dir()
app = Flask(
    __name__,
    template_folder=str(_BASE / "templates"),
    static_folder=str(_BASE / "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")
app.jinja_env.filters["mascara_cpf"] = _formatar_cpf
app.jinja_env.filters["mascara_telefone"] = _formatar_telefone


@app.route("/fotos/<path:nome>")
def foto(nome):
    return send_from_directory(str(UPLOAD_DIR), nome)


@app.context_processor
def _injetar_helpers():
    def foto_url(nome):
        """Resolve a URL da foto: se ja for uma URL do Supabase Storage
        (fotos novas), usa direto; se for so um nome de arquivo (fotos
        antigas, salvas localmente antes da migracao), busca na rota
        local /fotos/<nome>."""
        if not nome:
            return ""
        if nome.startswith("http://") or nome.startswith("https://"):
            return nome
        return url_for("foto", nome=nome)
    return dict(foto_url=foto_url, qtd_pendentes_offline=contar_pendentes())

# Campos considerados na checagem de "ficha completa". A foto fica de fora
# porque nem sempre é possível tirar foto do produtor na hora da visita.
FIELDS_OBRIGATORIOS = [f for f in FIELDS if f != "foto_produtor"]


def _campos_faltando(produtor) -> list:
    """Devolve a lista de campos obrigatorios que ainda estao vazios
    nesse cadastro."""
    faltando = []
    for f in FIELDS_OBRIGATORIOS:
        valor = produtor.get(f) or ""
        if not str(valor).strip():
            faltando.append(f)
    return faltando


@app.route("/dashboard")
def dashboard():
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM produtores ORDER BY nome_produtor"
    ).fetchall()
    conn.close()

    total = len(rows)
    completos = []
    incompletos = []
    for row in rows:
        produtor = dict(row)
        faltando = _campos_faltando(produtor)
        if faltando:
            incompletos.append({
                "id": produtor["id"],
                "nome_produtor": produtor.get("nome_produtor") or "-",
                "municipio": produtor.get("municipio") or "-",
                "tecnico_responsavel": produtor.get("tecnico_responsavel") or "-",
                "qtd_faltando": len(faltando),
                "labels_faltando": [FIELD_LABELS.get(f, f) for f in faltando],
            })
        else:
            completos.append(produtor)

    # Ordena quem tem mais campos faltando primeiro
    incompletos.sort(key=lambda p: p["qtd_faltando"], reverse=True)

    qtd_completos = len(completos)
    qtd_incompletos = len(incompletos)
    pct_completos = round((qtd_completos / total) * 100) if total else 0

    return render_template(
        "dashboard.html",
        total=total,
        qtd_completos=qtd_completos,
        qtd_incompletos=qtd_incompletos,
        pct_completos=pct_completos,
        incompletos=incompletos,
    )


@app.route("/")
def index():
    busca = request.args.get("q", "").strip()
    conn = get_connection()
    if busca:
        like = f"%{busca}%"
        rows = conn.execute(
            """SELECT * FROM produtores
               WHERE nome_produtor ILIKE ? OR cpf ILIKE ? OR municipio ILIKE ?
                  OR nome_propriedade ILIKE ?
               ORDER BY nome_produtor""",
            (like, like, like, like),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM produtores ORDER BY nome_produtor"
        ).fetchall()
    conn.close()
    return render_template("list.html", produtores=rows, busca=busca)


@app.route("/ficha/<int:pid>")
def ficha(pid):
    conn = get_connection()
    row = conn.execute("SELECT * FROM produtores WHERE id = ?", (pid,)).fetchone()
    conn.close()
    if row is None:
        abort(404)
    produtor = dict(row)
    produtor["data_nascimento"] = _exibir_data(produtor.get("data_nascimento"))
    produtor["cpf"] = _formatar_cpf(produtor.get("cpf"))
    produtor["telefone"] = _formatar_telefone(produtor.get("telefone"))
    return render_template("ficha.html", produtor=produtor)


@app.route("/novo", methods=["GET", "POST"])
def novo():
    if request.method == "POST":
        ok, dados = _salvar(None)
        if ok:
            return redirect(url_for("index"))
        return render_template("form.html", produtor=dados, municipios=MUNICIPIOS_CEARA)
    return render_template("form.html", produtor=None, municipios=MUNICIPIOS_CEARA)


@app.route("/editar/<int:pid>", methods=["GET", "POST"])
def editar(pid):
    conn = get_connection()
    row = conn.execute("SELECT * FROM produtores WHERE id = ?", (pid,)).fetchone()
    conn.close()
    if row is None:
        abort(404)
    if request.method == "POST":
        ok, dados = _salvar(pid)
        if ok:
            return redirect(url_for("index"))
        return render_template("form.html", produtor=dados, municipios=MUNICIPIOS_CEARA)
    produtor = dict(row)
    produtor["data_nascimento"] = _exibir_data(produtor.get("data_nascimento"))
    produtor["cpf"] = _formatar_cpf(produtor.get("cpf"))
    produtor["telefone"] = _formatar_telefone(produtor.get("telefone"))
    return render_template("form.html", produtor=produtor, municipios=MUNICIPIOS_CEARA)


def _erro_de_conexao(e: Exception) -> bool:
    """Detecta se a excecao e por falta de internet/conexao com o Supabase
    (para diferenciar de outros bugs de verdade, que devem continuar
    aparecendo normalmente como erro)."""
    if isinstance(e, (psycopg2.OperationalError, requests.exceptions.ConnectionError, requests.exceptions.Timeout)):
        return True
    # psycopg2 as vezes embrulha o erro de conexao dentro de outro tipo -
    # olhamos tambem o texto da mensagem, por seguranca.
    texto = str(e).lower()
    return any(p in texto for p in [
        "could not connect", "connection refused", "network is unreachable",
        "timeout expired", "temporary failure in name resolution",
        "failed to establish a new connection",
    ])


def _salvar(pid):
    # A foto e lida uma unica vez aqui (antes de mais nada), para podermos
    # tanto enviar ao Supabase Storage quanto, se precisar, guardar na fila
    # offline local.
    arquivo_foto = request.files.get("foto_produtor")
    foto_bytes = None
    foto_nome_original = ""
    foto_mimetype = ""
    if arquivo_foto and arquivo_foto.filename:
        foto_bytes = arquivo_foto.read()
        foto_nome_original = arquivo_foto.filename
        foto_mimetype = arquivo_foto.mimetype or ""

    dados = {}
    for f in FIELDS:
        if f == "foto_produtor":
            continue  # tratada a parte, acima
        valor = request.form.get(f, "").strip()
        if f in CAMPOS_MAIUSCULOS:
            valor = valor.upper()
        elif f == "cpf":
            valor = _formatar_cpf(valor)
        elif f == "telefone":
            valor = _formatar_telefone(valor)
        elif f in CAMPOS_SOMENTE_NUMERO:
            valor = re.sub(r"\D", "", valor)
        elif f in CAMPOS_DECIMAL:
            valor = re.sub(r"[^0-9,]", "", valor)
        elif f == "data_nascimento":
            valor = _normalizar_data(valor)
        dados[f] = valor

    # Volume Diario e Vacas em Lactacao sao obrigatorios
    if not dados.get("volume_diario") or not dados.get("vacas_lactacao"):
        flash(
            "Preencha \"Volume Diário Total (L/dia)\" e \"Vacas em Lactação\" "
            "- eles são obrigatórios e usados para calcular a produtividade.",
            "error",
        )
        dados["foto_produtor"] = _foto_atual(pid)
        return False, dados

    # Produtividade Media = Volume Diario Total / Vacas em Lactacao
    # (recalculado aqui no servidor, garantindo que fique sempre correta
    # mesmo que o calculo automatico da tela nao tenha rodado)
    try:
        volume = float(dados["volume_diario"].replace(",", "."))
        vacas = int(dados["vacas_lactacao"])
        dados["produtividade_media"] = f"{volume / vacas:.2f}".replace(".", ",")
    except (ValueError, ZeroDivisionError):
        dados["produtividade_media"] = ""

    # ---- Tenta salvar direto no Supabase (banco + foto) ----
    try:
        foto_final = _foto_atual(pid)
        if foto_bytes:
            foto_final = enviar_bytes(foto_bytes, foto_nome_original, foto_mimetype)
        dados["foto_produtor"] = foto_final

        conn = get_connection()
        if pid is None:
            cols = ", ".join(FIELDS)
            placeholders = ", ".join(["?"] * len(FIELDS))
            conn.execute(
                f"INSERT INTO produtores ({cols}) VALUES ({placeholders})",
                [dados[f] for f in FIELDS],
            )
            flash("Produtor cadastrado com sucesso.", "success")
        else:
            set_clause = ", ".join([f"{f} = ?" for f in FIELDS])
            conn.execute(
                f"UPDATE produtores SET {set_clause}, atualizado_em = now() WHERE id = ?",
                [dados[f] for f in FIELDS] + [pid],
            )
            flash("Cadastro atualizado com sucesso.", "success")
        conn.commit()
        conn.close()
        return True, dados

    except Exception as e:
        if not _erro_de_conexao(e):
            raise  # erro de verdade (bug) - nao esconde, deixa aparecer

        if pid is not None:
            # Editar um cadastro que ja existe no servidor exige internet
            # (nao da pra "mesclar" edicoes offline com seguranca aqui).
            flash(
                "Sem conexão com a internet no momento - não é possível "
                "editar um cadastro que já existe no servidor enquanto "
                "estiver offline. Tente novamente quando tiver internet.",
                "error",
            )
            dados["foto_produtor"] = _foto_atual(pid)
            return False, dados

        # Cadastro novo -> guarda na fila local para sincronizar depois
        dados["foto_produtor"] = ""
        id_local = adicionar_na_fila(dados, foto_bytes, foto_nome_original, foto_mimetype)
        flash(
            f"Sem internet no momento — o cadastro foi salvo aqui no "
            f"computador (pendência #{id_local}) e será enviado "
            f"automaticamente quando você sincronizar.",
            "aviso",
        )
        return True, dados


@app.route("/fila")
def fila():
    pendentes = listar_pendentes()
    return render_template("fila.html", pendentes=pendentes)


@app.route("/sincronizar", methods=["POST"])
def sincronizar():
    pendentes = listar_pendentes()
    sucesso = 0
    falha = 0
    for item in pendentes:
        try:
            foto_url = ""
            if item.get("foto_local_arquivo"):
                caminho = FILA_FOTOS_DIR / item["foto_local_arquivo"]
                if caminho.exists():
                    foto_url = enviar_bytes(
                        caminho.read_bytes(), item["foto_local_arquivo"], ""
                    )

            dados = {f: (item.get(f) or "") for f in FIELDS if f != "foto_produtor"}
            dados["foto_produtor"] = foto_url

            conn = get_connection()
            cols = ", ".join(FIELDS)
            placeholders = ", ".join(["?"] * len(FIELDS))
            conn.execute(
                f"INSERT INTO produtores ({cols}) VALUES ({placeholders})",
                [dados[f] for f in FIELDS],
            )
            conn.commit()
            conn.close()
            remover_da_fila(item["id_local"])
            sucesso += 1
        except Exception:
            falha += 1

    if sucesso:
        flash(f"{sucesso} cadastro(s) sincronizado(s) com sucesso!", "success")
    if falha:
        flash(
            f"{falha} cadastro(s) ainda não puderam ser enviados (sem "
            f"internet?). Eles continuam guardados aqui, tente de novo mais tarde.",
            "error",
        )
    if not sucesso and not falha:
        flash("Não há cadastros pendentes para sincronizar.", "success")
    return redirect(url_for("fila"))


@app.route("/fila/excluir/<int:id_local>", methods=["POST"])
def excluir_pendente(id_local):
    remover_da_fila(id_local)
    flash("Pendência removida da fila.", "success")
    return redirect(url_for("fila"))


@app.route("/excluir/<int:pid>", methods=["POST"])
def excluir(pid):
    conn = get_connection()
    conn.execute("DELETE FROM produtores WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    flash("Cadastro excluído.", "success")
    return redirect(url_for("index"))


@app.route("/pdf/<int:pid>")
def pdf(pid):
    conn = get_connection()
    row = conn.execute("SELECT * FROM produtores WHERE id = ?", (pid,)).fetchone()
    conn.close()
    if row is None:
        abort(404)
    buf = generate_pdf(row)
    nome = (row["nome_produtor"] or "produtor").strip().replace(" ", "_")
    return send_file(
        buf, mimetype="application/pdf", as_attachment=False,
        download_name=f"FATS_{nome}.pdf",
    )


@app.route("/exportar-excel")
def exportar_excel():
    busca = request.args.get("q", "").strip()
    conn = get_connection()
    if busca:
        like = f"%{busca}%"
        rows = conn.execute(
            """SELECT * FROM produtores
               WHERE nome_produtor ILIKE ? OR cpf ILIKE ? OR municipio ILIKE ?
                  OR nome_propriedade ILIKE ?
               ORDER BY nome_produtor""",
            (like, like, like, like),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM produtores ORDER BY nome_produtor"
        ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtores"

    # Cabecalho: ID + todos os campos, na ordem da ficha
    headers = ["ID"] + [FIELD_LABELS.get(f, f) for f in FIELDS] + ["Criado em", "Atualizado em"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E2C", end_color="1F4E2C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row in rows:
        valores = []
        for f in FIELDS:
            v = row[f]
            if f in CAMPOS_MAIUSCULOS and v:
                v = v.upper()
            elif f == "data_nascimento":
                v = _exibir_data(v)
            elif f == "cpf":
                v = _formatar_cpf(v)
            elif f == "telefone":
                v = _formatar_telefone(v)
            valores.append(v)
        linha = [row["id"]] + valores + [row["criado_em"], row["atualizado_em"]]
        ws.append(linha)

    # Largura automatica (aproximada) das colunas
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = len(str(header))
        for row_cells in ws.iter_rows(min_col=i, max_col=i, min_row=2):
            valor = row_cells[0].value
            if valor is not None:
                max_len = max(max_len, len(str(valor)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="produtores_fats.xlsx",
    )


if __name__ == "__main__":
    init_db()
    if getattr(sys, "frozen", False):
        # Rodando como .exe: abre o navegador automaticamente e roda sem
        # o modo debug/reloader (que nao funciona dentro do PyInstaller).
        import threading
        import webbrowser

        url = "http://127.0.0.1:5000"
        print("=" * 50)
        print(" Sistema FATS iniciado!")
        print(f" Abrindo no navegador: {url}")
        print(" Para encerrar o programa, feche esta janela.")
        print("=" * 50)
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()
        app.run(debug=False, host="127.0.0.1", port=5000)
    else:
        app.run(debug=True, host="0.0.0.0", port=5000)
